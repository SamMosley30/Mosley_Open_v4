"""This file contains the leaderboard class and its associated functions."""

import openpyxl as xl
from openpyxl.styles import *
from os import startfile, path

from .active_player import ActiveTournament

class Leaderboard(xl.Workbook):
    """This class represents the properties of the mosley open leaderboard and the functions needed for each type of
    leaderboard."""

    def __init__(self, active_player_list, calcutta_teams):
        super().__init__()
        self.players = list(active_player_list.active_player_list.values())
        self.teams = list(calcutta_teams.values())

        self.mosley_open = self.active
        self.mosley_open.title = 'Mosley Open'
        self.twisted_creek = self.create_sheet("Twisted Creek Classic")
        self.calcutta = self.create_sheet("Calcutta")
        self.dogfight1 = self.create_sheet("Day 1 Dogfight")
        self.dogfight2 = self.create_sheet("Day 2 Dogfight")
        self.dogfight3 = self.create_sheet("Day 3 Dogfight")
        dogfights = [self.dogfight1, self.dogfight2, self.dogfight3]

        self.mosley_open_header = ['Player', 'Point Target', 'Total Score', 'Position']
        self.twisted_creek_header = ['Player', 'Point Target', 'Total Score', 'Position']
        self.dogfight_header = ['Player', 'Point Target', 'Points', 'Net Points', 'Position']
        self.calcutta_header = ['Team', 'Total', 'Position']

        mosley_open_data = self.generate_mosley_open()
        twisted_creek_data = self.generate_twisted_creek()
        calcutta_data = self.generate_calcutta()
        dogfight_data = [self.generate_dogfight(day) for i, day in enumerate(self.players[0].days)]

        format_sheet(self.mosley_open, 'Mosley Open', self.mosley_open_header, mosley_open_data)
        format_sheet(self.twisted_creek, 'Twisted Creek', self.twisted_creek_header, twisted_creek_data)
        format_sheet(self.calcutta, 'Calcutta', self.calcutta_header, calcutta_data)
        for i, data in enumerate(dogfight_data):
            format_sheet(dogfights[i], f'Day {i+1} Dogfight', self.dogfight_header, data)
        self.save("MosleyOpen.xlsx")

        file = path.abspath('MosleyOpen.xlsx')
        startfile(file)

    def generate_mosley_open(self):
        """This function will extract all of the Mosley Open relevant information for printing into the leaderboard."""

        mosley_open_players = [player for player in self.players if player.active_tournament != ActiveTournament.TwistedCreek]
        mosley_open_players = sorted(mosley_open_players, key=lambda player: (player.net_mosley_open_points, player.total_raw_score), reverse=True)

        data = []
        for i, player in enumerate(mosley_open_players):
            player_data = [player.player_info.name, 36-int(player.player_info.mosley_open_handicap)]
            for day in player.days.values():
                if i == 0:
                    self.mosley_open_header[-2:-2] = ['Points', 'Net Points']
                player_data.extend([int(day.points), int(day.get_net_points(player.player_info.mosley_open_handicap))])
            player_data.extend([int(player.net_mosley_open_points), i+1])
            data.append(player_data)

        return data

    def generate_twisted_creek(self):
        """This function will extract all of the Twisted Creek relevant information for printing
        into the leaderboard."""

        twisted_creek_players = [player for player in self.players if player.active_tournament != ActiveTournament.MosleyOpen]
        twisted_creek_players = sorted(twisted_creek_players, key=lambda player: (player.net_twisted_creek_score, player.total_raw_score), reverse=True)

        data = []
        for i, player in enumerate(twisted_creek_players):
            player_data = [player.player_info.name, 36-player.player_info.twisted_creek_handicap]
            for day in player.days.values():
                if i == 0:
                    self.twisted_creek_header[-2:-2] = ['Points', 'Net Points']
                player_data.extend([int(day.points), int(day.get_net_points(player.player_info.twisted_creek_handicap))])
            player_data.extend([player.net_twisted_creek_score, i+1])
            data.append(player_data)

        return data

    def generate_calcutta(self):

        self.teams.sort(key=lambda x: x.total_points, reverse=True)

        data = list()
        for i, team in enumerate(self.teams):
            for j, day in enumerate(self.players[0].days):
                if i == 0:
                    self.calcutta_header.insert(-2, f'Day {j+1}')

            team_data = list(
                filter(
                    None,
                    [f'{team.player_1} and {team.player_2}', str(team.days['Day 1']), str(team.days['Day 2']),
                     str(team.days['Day 3']), team.total_points, i+1]
                )
            )
            data.append(team_data)

        return data

    def generate_dogfight(self, day):

        sorted_players = sorted(self.players, key=lambda x: [x.get_twisted_creek_daily_points(day), x.days[day].points], reverse=True)
        data = list()
        for i, player in enumerate(sorted_players):
            points_needed = 36 - player.player_info.twisted_creek_handicap
            day_data = player.days[day]
            player_data = [player.player_info.name, points_needed, day_data.points, player.get_twisted_creek_daily_points(day), i+1]
            data.append(player_data)

        return data


def format_sheet(sheet, title, header, data):

    black_fill = PatternFill(start_color='00000000',
                             end_color='00000000',
                             fill_type='solid')
    gold_fill = PatternFill(start_color='BF8F00',
                            end_color='BF8F00',
                            fill_type='solid')
    grey_fill = PatternFill(start_color='BFBFBF',
                            end_color='BFBFBF',
                            fill_type='solid')

    table_columns = len(data[0])
    table_rows = len(data)+2
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table_columns)
    sheet['A1'] = title
    sheet.append(header)
    for entry in data:
        sheet.append(entry)

    for i, row_cells in enumerate(sheet.rows):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='center')
            if cell.value is title:
                cell.fill = black_fill
                cell.font = Font(name='Times New Roman', size=24, bold=True, color='FFFFFFFF')
            elif cell.value in header:
                cell.fill = black_fill
                cell.font = Font(name='Times New Roman', size=20, bold=True, color='FFFFFFFF')
            elif row_cells[-1].value == 1:
                cell.fill = gold_fill
                cell.font = Font(name='Times New Roman', size=18, color='00000000')
            else:
                cell.fill = grey_fill
                cell.font = Font(name='Times New Roman', size=18, color='00000000')
            thick = Side(border_style="thick", color="000000")
            thin = Side(border_style="thin", color="000000")

            left_border = thin
            right_border = thin
            top_border = thin
            bottom_border = thin
            if cell == row_cells[0]:
                left_border = thick
                right_border = thin
            elif cell == row_cells[-1]:
                left_border = thin
                right_border = thick
            if i == 0:
                top_border = thick
                bottom_border = thin
            elif i == table_rows-1:
                top_border = thin
                bottom_border = thick

            cell.border = Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[1].column_letter].width = 1.8 * length + 3


def as_text(value):
    if not value:
        return ""

    return str(value)